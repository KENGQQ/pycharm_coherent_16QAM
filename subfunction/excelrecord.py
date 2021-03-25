from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl
import datetime
import os


def open_excel(address):
    if not os.path.exists(address + 'Record.xlsx'):
        print('creat an excel file')
        wb = Workbook()
        ws = wb.create_sheet('trial_1')
        del wb['Sheet']

    else:
        wb = openpyxl.load_workbook(address + 'Record.xlsx')
        ws = wb.create_sheet('trial_' + str(len(wb.worksheets) + 1), 0)

    ws['A1'] = datetime.datetime.now()

    parameter = ['eyepos', 'CMA[mean, type, overhead, taps, \nstepsize, iterator, earlystop, step_adj]', 'CMA [costX,costY]','PLL BW','X_corr ',
                 ' X[SNR,EVM,BERcount]','Y_corr ', 'Y[SNR,EVM,BERcount]']

    ws.column_dimensions['A'].width = 20.0
    ws.column_dimensions['B'].width = 8.0
    ws.column_dimensions['C'].width = 60.0
    ws.column_dimensions['D'].width = 25.0
    ws.column_dimensions['E'].width = 10.0
    ws.column_dimensions['F'].width = 30.0
    ws.column_dimensions['G'].width = 25.0
    ws.column_dimensions['H'].width = 30.0
    ws.column_dimensions['I'].width = 25.0


    ws.row_dimensions[1].height = 30

    for i in range(len(parameter)):
        ws.cell(row=1, column=i + 2, value=parameter[i])

    wb.save(address + 'Record.xlsx')


def write_excel(address, parameter_record):
    wb = openpyxl.load_workbook(address + 'Record.xlsx')
    ws = wb.worksheets[0]
    maxrow = ws.max_row
    for i in range(len(parameter_record)):
        ws.cell(row=maxrow + 1, column=1, value=datetime.datetime.now())
        ws.cell(row=maxrow + 1, column=i + 2, value=parameter_record[i])

    wb.save(address + 'Record.xlsx')
